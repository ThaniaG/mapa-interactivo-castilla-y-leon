"""
generar_datos_salud.py
══════════════════════════════════════════════════════════════════
Genera datos_salud.js y datos_puntos_salud.js con la cobertura
sanitaria por municipio de Castilla y León.

Nivel de cobertura (0-3):
  3 → Hospital (CNH_2025.xlsx)
  2 → Sede de Centro de Salud
  1 → Consultorio Local
  0 → Solo asignado a CS de otra localidad

Fuentes (en orden de prioridad):
  - CNH_2025.xlsx                      → hospitales (45 en CyL)
  - 2026_C_Catal_Centros_AP.xlsx       → CS y consultorios (catálogo SNS 2026)
  - centros-de-salud-municipios.csv    → asignación CS y datos de zona (Sacyl)
  - dependencia-entre-consultorios-y-centros-de-salud.csv → consultorios adicionales
  - 2026_C_Catal_disp_urg.xlsx         → PAC / urgencias extrahospitalarias

Coordenadas:
  - Hospitales, CS y PAC: geocodificación Nominatim (dirección real), con caché
  - Consultorios: centroide del municipio (sin dirección disponible)

Cómo ejecutar:
    python generar_datos_salud.py
"""

import csv
import json
import os
import re
import time
import unicodedata
import urllib.request
import urllib.parse
import openpyxl

BASE   = os.path.dirname(os.path.abspath(__file__))
CS_DIR = os.path.join(BASE, 'centros de salud')

F_CS_MUNI  = os.path.join(CS_DIR, 'centros-de-salud-municipios.csv')
F_DEPEND   = os.path.join(CS_DIR, 'dependencia-entre-consultorios-y-centros-de-salud.csv')
F_CNH      = os.path.join(CS_DIR, 'CNH_2025.xlsx')
F_CATALOGO = os.path.join(CS_DIR, '2026_C_Catal_Centros_AP.xlsx')
F_URG      = os.path.join(CS_DIR, '2026_C_Catal_disp_urg.xlsx')
F_EDADES   = os.path.join(BASE, 'datos_edades.js')
F_GEO      = os.path.join(BASE, 'datos_municipios.js')
F_GEOCACHE = os.path.join(CS_DIR, 'geocode_cache.json')

SALIDA_VAR    = os.path.join(BASE, 'datos_salud.js')
SALIDA_PUNTOS = os.path.join(BASE, 'datos_puntos_salud.js')

CYL_PROV = {'05', '09', '24', '34', '37', '40', '42', '47', '49'}

PROV_CATALOG = {
    'AVILA': '05', 'BURGOS': '09', 'LEON': '24', 'PALENCIA': '34',
    'SALAMANCA': '37', 'SEGOVIA': '40', 'SORIA': '42', 'VALLADOLID': '47', 'ZAMORA': '49'
}


# ── Utilidades de normalización ───────────────────────────────────────────────

def norm(texto):
    t = str(texto).upper().strip()
    t = unicodedata.normalize('NFD', t)
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    t = re.sub(r'[^A-Z0-9 ]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()

def variantes(n):
    """Cubre: "LA BANEZA" <-> "BANEZA LA" <-> "BANEZA, LA" <-> "BANEZA" """
    v = [n]
    for art in ['LA', 'EL', 'LOS', 'LAS', 'LO']:
        art_sp = art + ' '
        if n.startswith(art_sp):
            sin = n[len(art_sp):]
            v += [f'{sin}, {art}', sin, f'{sin} {art}']
        elif f', {art}' in n:
            base = n.replace(f', {art}', '').strip()
            v += [f'{art_sp}{base}', base, f'{base} {art}']
        elif n.endswith(f' {art}'):
            base = n[:-len(art)-1].strip()
            v += [f'{art_sp}{base}', f'{base}, {art}', base]
    return list(dict.fromkeys(v))


# ── Geocodificación con Nominatim ─────────────────────────────────────────────

print('Cargando caché de geocodificación...')
if os.path.exists(F_GEOCACHE):
    with open(F_GEOCACHE, encoding='utf-8') as f:
        geocache = json.load(f)
else:
    geocache = {}
print(f'  {len(geocache)} entradas en caché')

def geocodificar(query):
    """Devuelve (lat, lon) para la query dada, usando caché. None si falla."""
    if query in geocache:
        return geocache[query]
    params = urllib.parse.urlencode({
        'q': query, 'format': 'json', 'limit': '1', 'countrycodes': 'es'
    })
    url = f'https://nominatim.openstreetmap.org/search?{params}'
    req = urllib.request.Request(url, headers={'User-Agent': 'CyLDashboard/1.0 academic'})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        result = (round(float(data[0]['lat']), 6), round(float(data[0]['lon']), 6)) if data else None
    except Exception:
        result = None
    geocache[query] = result
    time.sleep(1.1)   # respetar límite Nominatim: 1 petición/segundo
    return result

def guardar_cache():
    with open(F_GEOCACHE, 'w', encoding='utf-8') as f:
        json.dump(geocache, f, ensure_ascii=False)


# ── 1. Índice nombre normalizado → código INE ─────────────────────────────────

print('Cargando referencia INE...')
with open(F_EDADES, encoding='utf-8') as f:
    edades_js = f.read()
edades = json.loads(edades_js.split('const DATOS_EDADES = ')[1].rstrip(';\n'))

idx_ine = {}
for cod, d in edades.items():
    pref = cod[:2]
    for v in variantes(norm(d['nombre'])):
        idx_ine[(pref, v)] = cod
print(f'  {len(edades)} municipios en referencia INE')

PREF_PROV_NOMBRE = {
    'AVILA': '05', 'BURGOS': '09', 'LEON': '24', 'PALENCIA': '34',
    'SALAMANCA': '37', 'SEGOVIA': '40', 'SORIA': '42', 'VALLADOLID': '47', 'ZAMORA': '49'
}

def inferir_pref(nombre_gerencia):
    ng = norm(nombre_gerencia)
    for prov, pref in PREF_PROV_NOMBRE.items():
        if prov in ng:
            return pref
    return None

def buscar_ine(nombre_muni, pref):
    if not pref:
        return None
    for v in variantes(norm(nombre_muni)):
        cod = idx_ine.get((pref, v))
        if cod:
            return cod
    return None


# ── 2. Centroides de municipio desde el GeoJSON ───────────────────────────────

print('Calculando centroides...')
with open(F_GEO, encoding='utf-8') as f:
    geo_js = f.read()
geo = json.loads(geo_js.split('const DATOS_GEO = ')[1].rstrip(';\n'))

def centroide(ring):
    lats = [c[1] for c in ring]
    lons = [c[0] for c in ring]
    return sum(lats)/len(lats), sum(lons)/len(lons)

centroides = {}
for feat in geo['features']:
    cod = feat['properties']['codigo']
    geom = feat['geometry']
    rings = [geom['coordinates'][0]] if geom['type'] == 'Polygon' \
            else [p[0] for p in geom['coordinates']]
    pts = [centroide(r) for r in rings]
    centroides[cod] = (
        sum(p[0] for p in pts)/len(pts),
        sum(p[1] for p in pts)/len(pts)
    )
print(f'  {len(centroides)} centroides calculados')


# ── 3. Hospitales (CNH 2025) ──────────────────────────────────────────────────

print('Cargando hospitales...')
wb = openpyxl.load_workbook(F_CNH)
ws = wb.active
hdrs = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

hospitales_raw = []   # datos en bruto antes de geocodificar
muni_con_hospital = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(hdrs, row))
    prov = str(r.get('C\xf3d. Provincia') or '').zfill(2)
    if prov not in CYL_PROV:
        continue
    cod6 = str(r.get('C\xf3d. Municipio') or '').strip()
    if len(cod6) < 5:
        continue
    cod_ine  = cod6[:5]
    nombre_h = (r.get('Nombre Centro') or '').strip()
    muni_h   = (r.get('Municipio') or '').strip()
    camas    = r.get('CAMAS') or 0
    clase    = (r.get('Clase de Centro') or '').strip()
    direccion = (r.get('Direcci\xf3n') or '').strip()
    cp        = str(r.get('C\xf3digo Postal') or '').strip().zfill(5)

    if cod_ine not in muni_con_hospital:
        muni_con_hospital[cod_ine] = []
    muni_con_hospital[cod_ine].append({'nombre': nombre_h, 'clase': clase, 'camas': camas})

    hospitales_raw.append({
        'cod_ine': cod_ine,
        'tipo':    'hospital',
        'nombre':  nombre_h,
        'municipio': muni_h,
        'clase':   clase,
        'camas':   int(camas) if camas else 0,
        'dir_query': f'{direccion}, {cp}, {muni_h}, España' if direccion else None,
    })

print(f'  {len(muni_con_hospital)} municipios con hospital, {len(hospitales_raw)} registros')


# ── 4. Centros de Salud (CSV Sacyl) ───────────────────────────────────────────

print('Cargando centros de salud (CSV Sacyl)...')
with open(F_CS_MUNI, encoding='utf-8-sig') as f:
    cs_rows = list(csv.DictReader(f, delimiter=';'))

muni_a_cs    = {}
sedes_cs_ine = {}

for r in cs_rows:
    muni      = r.get('MUNICIPIO', '').strip()
    cs_name   = r.get('NOMBRE CENTRO SALUD', '').strip()
    zona      = r.get('NOMBRE ZONA', '').strip()
    gerencia  = r.get('NOMBRE GERENCIA', '').strip()
    cod_zona  = r.get('CÓDIGO ZONA', r.get('C\xd3DIGO ZONA', '')).strip()
    pac       = r.get('PAC', '').strip()

    pref = inferir_pref(gerencia)
    cod_ine = buscar_ine(muni, pref)
    if not cod_ine:
        continue

    muni_a_cs[cod_ine] = {
        'nombre_cs': cs_name, 'nombre_zona': zona,
        'cod_zona': cod_zona, 'pac': pac, 'gerencia': gerencia,
    }

    sede_raw   = re.sub(r'^C\.S\.\s+', '', cs_name).strip()
    datos_sede = {'nombre_cs': cs_name, 'nombre_zona': zona, 'pref': pref}
    cod_sede   = buscar_ine(sede_raw, pref) if pref else None
    if not cod_sede:
        for p in CYL_PROV:
            cod_sede = buscar_ine(sede_raw, p)
            if cod_sede:
                datos_sede['pref'] = p
                break
    if cod_sede:
        sedes_cs_ine[cod_sede] = datos_sede

print(f'  {len(muni_a_cs)} municipios con CS asignado')
print(f'  {len(sedes_cs_ine)} sedes CS resueltas (CSV Sacyl)')


# ── 4b. Catálogo SNS 2026: CS y Consultorios ─────────────────────────────────

print('Cargando catalogo SNS 2026...')
wb_cat = openpyxl.load_workbook(F_CATALOGO, read_only=True)
ws_cat = wb_cat['Catálogo - 2026']
cat_rows = list(ws_cat.iter_rows(values_only=True))

zona_a_sede = {}   # (pref, zona_norm) → (cod_ine, nombre_cs)
cat_cs_dir  = {}   # cod_ine → {dir_query}  para geocodificación

for r in cat_rows[1:]:
    if not r[0] or 'CASTILLA Y' not in str(r[0]) or 'MANCHA' in str(r[0]):
        continue
    if r[7] != 'CENTRO SALUD':
        continue
    prov_norm = norm(str(r[1]))
    pref = PROV_CATALOG.get(prov_norm)
    if not pref:
        continue
    zona      = str(r[3] or '').strip()
    muni      = str(r[4] or '').strip()
    nombre_cs = str(r[8] or '').strip()
    direccion = str(r[9] or '').strip()
    cp        = str(r[10] or '').strip().zfill(5)
    localidad = str(r[11] or '').strip()

    cod_ine = buscar_ine(muni, pref)
    if cod_ine:
        zona_a_sede[(pref, norm(zona))] = (cod_ine, nombre_cs)
        if direccion:
            cat_cs_dir[cod_ine] = f'{direccion}, {cp}, {localidad}, España'

print(f'  {len(zona_a_sede)} zonas con CS sede resueltas')

cat_sedes_nuevas = 0
for (pref, zona_norm), (cod_ine, nombre_cs) in zona_a_sede.items():
    if cod_ine not in sedes_cs_ine:
        sedes_cs_ine[cod_ine] = {'nombre_cs': nombre_cs, 'nombre_zona': zona_norm, 'pref': pref}
        cat_sedes_nuevas += 1

print(f'  {cat_sedes_nuevas} sedes CS nuevas desde catalogo, total: {len(sedes_cs_ine)}')

cat_consultorios = {}
for r in cat_rows[1:]:
    if not r[0] or 'CASTILLA Y' not in str(r[0]) or 'MANCHA' in str(r[0]):
        continue
    if r[7] != 'CONSULTORIO LOCAL':
        continue
    prov_norm = norm(str(r[1]))
    pref = PROV_CATALOG.get(prov_norm)
    if not pref:
        continue
    zona  = str(r[3] or '').strip()
    muni  = str(r[4] or '').strip()
    cod_ine = buscar_ine(muni, pref)
    if not cod_ine:
        continue
    cs_padre = zona_a_sede.get((pref, norm(zona)), (None, ''))[1]
    cat_consultorios[cod_ine] = cs_padre

print(f'  {len(cat_consultorios)} municipios con consultorio (catalogo)')


# ── 5. Consultorios adicionales (CSV dependencia) ─────────────────────────────

print('Cargando consultorios (CSV dependencia)...')
with open(F_DEPEND, encoding='utf-8-sig') as f:
    dep_rows = list(csv.DictReader(f, delimiter=';'))

muni_con_consultorio = dict(cat_consultorios)

for r in dep_rows:
    consultorio = r.get('CONSULTORIO', '').strip()
    cs_padre    = r.get('CENTRO', '').strip()
    gerencia    = r.get('NOMBRE GERENCIA', '').strip()
    if not consultorio.upper().startswith('C.L.'):
        continue
    muni_raw = re.sub(r'^C\.L\.\s+', '', consultorio).strip()
    pref = inferir_pref(gerencia)
    cod_ine = buscar_ine(muni_raw, pref)
    if not cod_ine:
        for p in CYL_PROV:
            cod_ine = buscar_ine(muni_raw, p)
            if cod_ine:
                break
    if cod_ine and cod_ine not in muni_con_consultorio:
        muni_con_consultorio[cod_ine] = cs_padre

print(f'  {len(muni_con_consultorio)} municipios con consultorio (total)')


# ── 5b. PAC / Urgencias Extrahospitalarias ────────────────────────────────────

print('Cargando PAC / urgencias extrahospitalarias...')
wb_urg = openpyxl.load_workbook(F_URG, read_only=True)
ws_urg = wb_urg['Catálogo - 2026']
urg_rows = list(ws_urg.iter_rows(values_only=True))

pac_raw = []   # datos en bruto antes de geocodificar

for r in urg_rows[1:]:
    if not r[0] or 'CASTILLA Y' not in str(r[0]) or 'MANCHA' in str(r[0]):
        continue
    prov_norm = norm(str(r[1]))
    pref = PROV_CATALOG.get(prov_norm)
    if not pref:
        continue
    muni        = str(r[2] or '').strip()
    tipo_centro = str(r[4] or '').strip()
    ubicacion   = str(r[5] or '').strip()
    direccion   = str(r[6] or '').strip()
    cp          = str(r[7] or '').strip().zfill(5)
    localidad   = str(r[8] or '').strip()

    cod_ine = buscar_ine(muni, pref)
    if not cod_ine or cod_ine not in centroides:
        continue

    tipo_key = 'cg' if 'CENTRO DE GUARDIA' in tipo_centro.upper() else 'pac'
    pac_raw.append({
        'cod_ine':   cod_ine,
        'tipo':      tipo_key,
        'nombre':    tipo_centro,
        'municipio': muni,
        'ubicacion': ubicacion,
        'dir_query': f'{direccion}, {cp}, {localidad}, España' if direccion else None,
    })

print(f'  {len(pac_raw)} puntos PAC/CG cargados')


# ── 6. Geocodificación de hospitales, CS y PAC ────────────────────────────────

def coords(cod_ine, dir_query):
    """Intenta geocodificar; si falla usa centroide."""
    if dir_query:
        geo_res = geocodificar(dir_query)
        if geo_res:
            return geo_res
    if cod_ine in centroides:
        lat, lon = centroides[cod_ine]
        return round(lat, 6), round(lon, 6)
    return None, None

# Contar cuántas peticiones nuevas haremos
nuevas_hospital = sum(1 for h in hospitales_raw if h['dir_query'] and h['dir_query'] not in geocache)
nuevas_cs       = sum(1 for c, d in cat_cs_dir.items() if d not in geocache)
nuevas_pac      = sum(1 for p in pac_raw if p['dir_query'] and p['dir_query'] not in geocache)
total_nuevas    = nuevas_hospital + nuevas_cs + nuevas_pac

print(f'\nGeocod. necesarias: {nuevas_hospital} hospitales + {nuevas_cs} CS + {nuevas_pac} PAC = {total_nuevas} peticiones')
if total_nuevas:
    print(f'  (tiempo estimado: ~{total_nuevas} segundos)')

# Hospitales
print('Geocodificando hospitales...')
hospitales = []
geo_h_ok = 0
for h in hospitales_raw:
    lat, lon = coords(h['cod_ine'], h['dir_query'])
    if lat is None:
        continue
    if h['dir_query'] and geocache.get(h['dir_query']):
        geo_h_ok += 1
    hospitales.append({
        'lat': lat, 'lon': lon,
        'tipo': 'hospital', 'nombre': h['nombre'],
        'municipio': h['municipio'], 'clase': h['clase'], 'camas': h['camas'],
    })
print(f'  {geo_h_ok}/{len(hospitales)} con coordenadas geocodificadas, resto centroide')

# CS
print('Geocodificando centros de salud...')
puntos_cs = []
geo_cs_ok = 0
for cod_ine, datos in sedes_cs_ine.items():
    dir_q = cat_cs_dir.get(cod_ine)
    lat, lon = coords(cod_ine, dir_q)
    if lat is None:
        continue
    if dir_q and geocache.get(dir_q):
        geo_cs_ok += 1
    nombre_muni = edades.get(cod_ine, {}).get('nombre', '')
    puntos_cs.append({
        'lat': lat, 'lon': lon,
        'tipo': 'cs', 'nombre': datos['nombre_cs'],
        'municipio': nombre_muni, 'zona': datos['nombre_zona'],
    })
print(f'  {geo_cs_ok}/{len(puntos_cs)} con coordenadas geocodificadas, resto centroide')

# Consultorios (siempre centroide)
puntos_consultorio = []
for cod_ine, cs_padre in muni_con_consultorio.items():
    if cod_ine in sedes_cs_ine or cod_ine in muni_con_hospital:
        continue
    if cod_ine not in centroides:
        continue
    lat, lon = centroides[cod_ine]
    nombre_muni = edades.get(cod_ine, {}).get('nombre', '')
    puntos_consultorio.append({
        'lat': round(lat, 6), 'lon': round(lon, 6),
        'tipo': 'consultorio',
        'nombre': f'C.L. {nombre_muni.upper()}',
        'municipio': nombre_muni, 'cs_padre': cs_padre,
    })

# PAC
print('Geocodificando PAC...')
puntos_pac = []
geo_pac_ok = 0
for p in pac_raw:
    lat, lon = coords(p['cod_ine'], p['dir_query'])
    if lat is None:
        continue
    if p['dir_query'] and geocache.get(p['dir_query']):
        geo_pac_ok += 1
    puntos_pac.append({
        'lat': lat, 'lon': lon,
        'tipo': p['tipo'], 'nombre': p['nombre'],
        'municipio': p['municipio'], 'ubicacion': p['ubicacion'],
    })
print(f'  {geo_pac_ok}/{len(puntos_pac)} con coordenadas geocodificadas, resto centroide')

guardar_cache()
print(f'Cache guardada: {len(geocache)} entradas')

print(f'\nResumen puntos:')
print(f'  Hospitales   : {len(hospitales)}')
print(f'  CS           : {len(puntos_cs)}')
print(f'  Consultorios : {len(puntos_consultorio)}')
print(f'  PAC/CG       : {len(puntos_pac)}')


# ── 7. Construir objeto datos_salud ───────────────────────────────────────────

print('\nConstruyendo datos_salud...')
resultado = {}

for cod_ine in edades:
    nivel = 0
    tipo  = 'sin_datos'

    cs_info = muni_a_cs.get(cod_ine, {})
    if not cs_info.get('nombre_cs') and cod_ine in cat_consultorios:
        cs_info = {'nombre_cs': cat_consultorios[cod_ine], 'nombre_zona': '', 'pac': ''}
    if not cs_info.get('nombre_cs') and cod_ine in sedes_cs_ine:
        cs_info = {'nombre_cs': sedes_cs_ine[cod_ine]['nombre_cs'],
                   'nombre_zona': sedes_cs_ine[cod_ine]['nombre_zona'], 'pac': ''}

    if cod_ine in muni_con_hospital:
        nivel = 3
        tipo  = 'hospital'
    elif cod_ine in sedes_cs_ine:
        nivel = 2
        tipo  = 'centro_salud'
    elif cod_ine in muni_con_consultorio:
        nivel = 1
        tipo  = 'consultorio'
    elif cod_ine in muni_a_cs:
        nivel = 0
        tipo  = 'asignado'

    resultado[cod_ine] = {
        'nombre':      edades[cod_ine]['nombre'],
        'nivel_salud': nivel,
        'tipo':        tipo,
        'nombre_cs':   cs_info.get('nombre_cs', ''),
        'nombre_zona': cs_info.get('nombre_zona', ''),
        'pac':         cs_info.get('pac', ''),
    }

from collections import Counter
dist = Counter(v['tipo'] for v in resultado.values())
print(f'  hospital      : {dist["hospital"]}')
print(f'  centro_salud  : {dist["centro_salud"]}')
print(f'  consultorio   : {dist["consultorio"]}')
print(f'  asignado      : {dist["asignado"]}')
print(f'  sin_datos     : {dist["sin_datos"]}')


# ── 7b. Puntos para municipios sin centro propio ──────────────────────────────
# Asignados (municipio que depende del CS de otra localidad)
# Sin_datos (municipio sin información sanitaria disponible)

puntos_sin_centro = []
for cod_ine, v in resultado.items():
    if v['tipo'] not in ('asignado', 'sin_datos'):
        continue
    if cod_ine not in centroides:
        continue
    lat, lon = centroides[cod_ine]
    puntos_sin_centro.append({
        'lat':      round(lat, 6),
        'lon':      round(lon, 6),
        'tipo':     v['tipo'],
        'nombre':   v['nombre'],
        'municipio': v['nombre'],
        'nombre_cs': v.get('nombre_cs', ''),
    })

print(f'  Asignados + sin_datos como puntos: {len(puntos_sin_centro)}')


# ── 8. Guardar datos_salud.js ─────────────────────────────────────────────────

resultado = dict(sorted(resultado.items()))
json_str = json.dumps(resultado, ensure_ascii=False, separators=(',', ':'))
with open(SALIDA_VAR, 'w', encoding='utf-8') as f:
    f.write('// datos_salud.js — Generado por generar_datos_salud.py\n')
    f.write('// Cobertura sanitaria por municipio de Castilla y León\n')
    f.write('// Nivel: 0=asignado, 1=consultorio, 2=centro salud, 3=hospital\n')
    f.write('// Fuentes: CNH_2025 + Cat.SNS 2026 + Sacyl\n')
    f.write('// NO editar manualmente.\n\n')
    f.write(f'const DATOS_SALUD = {json_str};\n')
print(f'\nGenerado: {SALIDA_VAR}')


# ── 9. Guardar datos_puntos_salud.js ──────────────────────────────────────────

todos_puntos = hospitales + puntos_cs + puntos_consultorio + puntos_pac + puntos_sin_centro
json_pts = json.dumps(todos_puntos, ensure_ascii=False, separators=(',', ':'))
with open(SALIDA_PUNTOS, 'w', encoding='utf-8') as f:
    f.write('// datos_puntos_salud.js — Generado por generar_datos_salud.py\n')
    f.write(f'// {len(hospitales)} hosp · {len(puntos_cs)} CS · {len(puntos_consultorio)} CL · {len(puntos_pac)} PAC · {len(puntos_sin_centro)} sin_centro\n')
    f.write('// Coordenadas: Nominatim (hospitales/CS/PAC) + centroide (consultorios/asignados)\n')
    f.write('// NO editar manualmente.\n\n')
    f.write(f'const datosPuntosSalud = {json_pts};\n')
print(f'Generado: {SALIDA_PUNTOS}')
print(f'\nTotal puntos de salud: {len(todos_puntos):,}')
